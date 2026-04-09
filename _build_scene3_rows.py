# -*- coding: utf-8 -*-
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

base = r"c:\Users\林靖怡\Desktop\数据\场景三"
xlsx = os.path.join(base, "数据表格.xlsx")
out_js = os.path.join(os.path.dirname(__file__), "scene3_rows.json")

wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows_iter = ws.iter_rows(values_only=True)
header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
objs = []
for row in rows_iter:
    d = {}
    for i, h in enumerate(header):
        v = row[i] if i < len(row) else None
        if h in ("视觉复杂度", "文字字数", "包装程度"):
            try:
                d[h] = int(float(v)) if v not in (None, "") else 0
            except (ValueError, TypeError):
                d[h] = 0
        elif h == "贴近真实情绪程度指数":
            try:
                d[h] = float(v) if v not in (None, "") else None
            except (ValueError, TypeError):
                d[h] = None
        else:
            d[h] = "" if v is None else str(v).strip()
    objs.append(d)
wb.close()

with open(out_js, "w", encoding="utf-8") as f:
    json.dump(objs, f, ensure_ascii=False)

print(len(objs), "rows ->", out_js)
