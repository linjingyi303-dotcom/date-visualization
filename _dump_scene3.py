# -*- coding: utf-8 -*-
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

base = r"c:\Users\林靖怡\Desktop\数据\场景三"
xlsx = os.path.join(base, "数据表格.xlsx")
out_path = os.path.join(os.path.dirname(__file__), "scene3_sheet_dump.json")

wb = load_workbook(xlsx, read_only=True, data_only=True)
out = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append([str(c) if c is not None else "" for c in row])
        if i >= 500:
            break
    out[sn] = rows
wb.close()

with open(out_path, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=0)

print("OK", out_path)
for sn, rows in out.items():
    print("---", sn, len(rows))
    for r in rows[:4]:
        print(r[:18])
